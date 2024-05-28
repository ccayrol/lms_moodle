import csv
import os
import shutil
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from verification import Verification

# Une fonction qui crée deux dossiers appelés Dossier_Fiche_de_liaison_date_heure_minutes 
# et Dossier_Alternant_date_heure_minutes à la racine du projet
# où _date_heure_minutes est la date et l'heure de la création du dossier ainsi que les minutes
def creer_dossier():
    # Récupérer la date et l'heure actuelle
    date = datetime.now()
    date_heure = date.strftime("%Y-%m-%d_%H-%M")
    # Créer les noms des dossiers
    dossier_fiche_de_liaison = f"Dossier_Fiche_de_liaison_{date_heure}"
    dossier_alternant = f"Dossier_Alternant_{date_heure}"
    # Créer les dossiers
    os.mkdir(dossier_fiche_de_liaison)
    os.mkdir(dossier_alternant)
    # Retourner les noms des dossiers créés
    return dossier_fiche_de_liaison, dossier_alternant

# Fonction qui duplique un fichier modèle pour chaque étudiant
def dupliquer_fichier(fichier_base, nom_etudiant, nom_dossier_alternant):
    # Créer le chemin complet pour le nouveau fichier
    nom_fichier = f"{nom_dossier_alternant}\\Fiche_de_liaison_{nom_etudiant}.xlsx"
    # Vérifier si le fichier n'existe pas déjà
    if not os.path.exists(nom_fichier):
        # Copier le fichier de base vers le nouveau fichier
        shutil.copy(fichier_base, nom_fichier)
    return nom_fichier

# Fonction qui lit les données d'un fichier CSV dans un DataFrame pandas
def read_csv_data(csv_file, dtype):
    """Lire les données d'un fichier CSV dans un DataFrame pandas."""
    return pd.read_csv(csv_file, dtype=dtype)

# Fonction qui crée et remplit des fichiers Excel à partir des données CSV
def create_excel_file(excel_file, data, cell_mapping, sheet_name, nom_dossier_alternant):
    error_report = []

    # Parcourir les lignes du DataFrame
    for index, row in data.iterrows():
        # Charger le fichier Excel
        nom_fichier = dupliquer_fichier(excel_file, row["Nom complet"], nom_dossier_alternant)
        workbook = load_workbook(nom_fichier)

        # Vérifier si la feuille spécifiée existe
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name} does not exist in the workbook.")
        sheet = workbook[sheet_name]
        error = []
        # Remplir les cellules du fichier Excel avec les données du DataFrame
        for column, cell in cell_mapping.items():
            if not verif_champ(row[column], column):
                msg_error = f"Donnée invalide  {column} : {row[column]}, "
                error.append(msg_error)
            sheet[cell] = row[column]

        # Ajouter les erreurs à la liste de rapports d'erreur
        error_report.append({
            "Nom": row["Nom complet"],
            "Email": row["Nom d'utilisateur"],
            "error": error})

        # Sauvegarder le fichier Excel
        workbook.save(nom_fichier)
    # Écrire le rapport d'erreur dans un fichier CSV
    filename = f"{nom_dossier_alternant}\\error_report.csv"
    write_error_report_to_csv(error_report, filename)

# Fonction qui écrit un rapport d'erreur dans un fichier CSV
def write_error_report_to_csv(error_report, filename):
    # Les noms des champs pour le fichier CSV
    fieldnames = ["Nom", "Email", "Erreurs"]

    # Dictionnaire pour stocker les rapports d'erreur actuels
    existing_reports = {}

    # Lire le fichier existant si disponible
    if os.path.exists(filename):
        with open(filename, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                # Ajouter les lignes existantes au dictionnaire
                existing_reports[row["Email"]] = {
                    "Nom": row["Nom"],
                    "Email": row["Email"],
                    "Erreurs": row["Erreurs"].split(', ')  # Convertir la chaîne en liste
                }

    # Mettre à jour ou ajouter les rapports d'erreur
    for report in error_report:
        email = report["Email"]
        if email in existing_reports:
            # Remplacer les erreurs existantes par les nouvelles
            existing_reports[email]["Erreurs"] = report["error"]
        else:
            # Ajouter un nouveau rapport
            existing_reports[email] = {
                "Nom": report["Nom"],
                "Email": email,
                "Erreurs": report["error"]
            }

    reports_to_write = [report for report in existing_reports.values() if report["Erreurs"]]
    # Écrire les données mises à jour dans le fichier
    with open(filename, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)

        # Écrire l'en-tête
        writer.writeheader()

        # Écrire chaque ligne de rapport d'erreur
        for report in reports_to_write:
            writer.writerow({
                "Nom": report["Nom"],
                "Email": report["Email"],
                "Erreurs": ', '.join(report["Erreurs"])  # Convertir la liste d'erreurs en une chaîne
            })

# Fonction qui vérifie la validité des champs
def verif_champ(value, column):
    if column.startswith("Q01_Prenom") or column.startswith("Q11_Prenom") or column.startswith("Q17_Prenom"):
        return Verification.verify_prenom(value)
    elif column.startswith("Q02_Nom") or column.startswith("Q12_Nom") or column.startswith("Q18_Nom"):
        return Verification.verify_nom(value)
    elif column == "Q03_TelPortable" or column == "Q21_TelephonePortable":
        return Verification.verify_numero_telephone_france(value)
    elif column.startswith("Q04_Email") or column.startswith("Q15_Email") or column.startswith("Q20_Email"):
        return Verification.verify_email_etudiant(value) or Verification.verify_email_personnel(value)
    elif column == "Q07_NumeroSiret":
        return Verification.verify_siren_ou_siret_api(value)
    elif column == "Q08_CodeAPE":
        return True
    elif column == "Q09_CodeIDCC":
        return True
    elif column == "Q10_OPCOEntreprise":
        return True
    elif column == "Q14_Telephone" or column == "Q19_TelephoneFixe":
        return Verification.verify_numero_telephone_france(value)
    elif column == "Q16_Fonction":
        return True
    elif column == "Q23_DateDebutContrat":
        return Verification.verify_date_validite(value)
    elif column == "Q24_DateFinContrat":
        return Verification.verify_date_validite(value)
    elif column == "Q25_ServiceAccueillant":
        return True
    elif column == "Q26_IntituleDuPoste":
        return True
    elif column == "Q28_HorairesHebdomadaires":
        return True
    elif column == "Q29_AutresRemarques":
        return True
    else:
        return True

# Fonction principale qui remplit les fichiers Excel à partir des données CSV
def remplir_fichier_excel(nom_dossier_alternant):
    csv_file = 'Remplissage_Fiche_de_Liaison_Alternant.csv'  # Le fichier téléchargé depuis Moodle
    excel_file = 'Fiche_de_liaison_Alternant.xlsx'  # Le modèle à remplir
    sheet_name = 'CFA - Promesse Recrutement'  # La feuille du modèle à remplir

    # Mapping des colonnes du DataFrame aux cellules Excel
    cell_mapping = {
        'Q01_Prenom': 'C19',
        "Q02_Nom": 'F19',
        "Q03_TelPortable": 'C21',
        "Q04_Email": 'F21',
        "Q05_NomEntreprise": 'D26',
        "Q06_AdresseEntreprise": 'C28',
        "Q07_NumeroSiret": 'C31',
        "Q08_CodeAPE": 'G31',
        "Q09_CodeIDCC": 'F33',
        "Q10_OPCOEntreprise": 'F35',
        "Q11_Prenom": 'C40',
        "Q12_Nom": 'F40',
        "Q13_AdressePostale": 'C42',
        "Q14_Telephone": 'C45',
        "Q15_Email": 'F45',
        "Q16_Fonction": 'C50',
        "Q17_Prenom": 'C52',
        "Q18_Nom": 'F52',
        "Q19_TelephoneFixe": 'C54',
        "Q20_Email": 'F54',
        "Q21_TelephonePortable": 'C56',
        "Q22_Adresse": 'B59',
        "Q23_DateDebutContrat": 'D69',
        "Q24_DateFinContrat": 'D71',
        "Q25_ServiceAccueillant": 'D76',
        "Q26_IntituleDuPoste": 'B79',
        "Q27_Adresse": 'B102',
        "Q28_HorairesHebdomadaires": 'B104',
        "Q29_AutresRemarques": 'B115'
    }

    dtype = {'Q03_TelPortable': str,
             'Q07_NumeroSiret': str,
             'Q14_Telephone': str,
             'Q19_TelephoneFixe': str,
             'Q21_TelephonePortable': str,
             'Q07_NumeroSiret': str
             }
    
    # Lire les données du fichier CSV
    data = read_csv_data(csv_file, dtype)

    # Créer les fiches de liaison alternant pour chaque étudiant
    create_excel_file(excel_file, data, cell_mapping, sheet_name, nom_dossier_alternant)
